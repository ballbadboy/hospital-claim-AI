"""Report engine — generates Excel reports for hospital claim KPIs."""

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _safe_cell(value) -> str:
    """Prevent Excel formula injection by escaping values that start with formula chars."""
    s = str(value) if value is not None else ""
    if s and s[0] in ("=", "+", "-", "@", "|", "%"):
        return "\t" + s
    return s


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None and hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


class ReportEngine:
    @staticmethod
    def generate_monthly_summary(stats: dict) -> bytes:
        """Generate monthly summary Excel report."""
        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "สรุปรายเดือน"
        period = stats.get("period", "N/A")
        ws.append(["Hospital Claim AI — สรุปรายเดือน", "", period])
        ws.merge_cells("A1:C1")
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([])

        headers = ["ตัวชี้วัด", "ค่า"]
        ws.append(headers)
        _style_header_row(ws, 3, 2)

        ws.append(["จำนวนเคสทั้งหมด", stats.get("total_claims", 0)])
        ws.append(["เคสถูก Deny", stats.get("denied_claims", 0)])
        ws.append(["Deny Rate (%)", stats.get("deny_rate", 0)])
        ws.append(["Revenue at Risk (฿)", f"{stats.get('revenue_at_risk', 0):,.0f}"])
        ws.append(["Revenue Recovered (฿)", f"{stats.get('revenue_recovered', 0):,.0f}"])
        ws.append(["Average Score", stats.get("avg_score", 0)])
        _auto_width(ws)

        # Department sheet
        if stats.get("by_department"):
            ws2 = wb.create_sheet("แยกตามแผนก")
            ws2.append(["แผนก", "จำนวนเคส", "ถูก Deny", "Deny Rate (%)"])
            _style_header_row(ws2, 1, 4)
            for dept, data in stats["by_department"].items():
                ws2.append([dept, data.get("total", 0), data.get("denied", 0), data.get("deny_rate", 0)])
            _auto_width(ws2)

        # Deny reasons sheet
        if stats.get("by_deny_reason"):
            ws3 = wb.create_sheet("สาเหตุ Deny")
            ws3.append(["สาเหตุ", "จำนวน"])
            _style_header_row(ws3, 1, 2)
            for reason, count in stats["by_deny_reason"].items():
                ws3.append([reason, count])
            _auto_width(ws3)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_audit_trail(trail: list[dict], an: str) -> bytes:
        """Generate audit trail Excel for a specific claim."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Audit Trail - {an}"
        ws.append([f"Audit Trail: {an}"])
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([])

        ws.append(["Action", "User", "Details", "Timestamp"])
        _style_header_row(ws, 3, 4)

        for entry in trail:
            ws.append([
                _safe_cell(entry.get("action", "")),
                _safe_cell(entry.get("user", "")),
                _safe_cell(str(entry.get("details", {}))),
                _safe_cell(str(entry.get("at", ""))),
            ])
        _auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_department_detail(dept_data: dict) -> bytes:
        """Generate department detail report."""
        wb = Workbook()
        ws = wb.active
        dept = dept_data.get("department", "Unknown")
        period = dept_data.get("period", "N/A")
        ws.title = f"{dept} - {period}"

        ws.append([f"รายงานแผนก {dept} — {period}"])
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([])

        ws.append(["ตัวชี้วัด", "ค่า"])
        _style_header_row(ws, 3, 2)
        ws.append(["จำนวนเคส", dept_data.get("total", 0)])
        ws.append(["ถูก Deny", dept_data.get("denied", 0)])
        ws.append(["Deny Rate (%)", dept_data.get("deny_rate", 0)])
        ws.append(["Average Score", dept_data.get("avg_score", 0)])

        if dept_data.get("top_deny_reasons"):
            ws.append([])
            ws.append(["สาเหตุ Deny หลัก", "จำนวน"])
            row_num = ws.max_row
            _style_header_row(ws, row_num, 2)
            for reason, count in dept_data["top_deny_reasons"]:
                ws.append([reason, count])

        _auto_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
